import requests # type: ignore
from bs4 import BeautifulSoup # type: ignore
from selenium import  # type: ignore
from selenium.webdriver.common.by import By # type: ignore
from selenium.webdriver.support.ui import WebDriverWait # type: ignore
from selenium.webdriver.support import expected_conditions as EC# type: ignore
from selenium.common.exceptions import TimeoutException, NoSuchElementException# type: ignore
import pandas as pd# type: ignore
import time
import logging
import json
import os
from uuid import uuid4
import re
from datetime import datetime
from openpyxl import Workbook # type: ignore
from openpyxl.styles import PatternFill, Font, Alignment # type: ignore
from openpyxl.utils import get_column_letter # type: ignore


class AdvancedVapeScraper:
    def __init__(self, job_id=None):
        self.job_id = job_id or str(uuid4())
        self.driver = None
        self.products_data = []
        self.is_running = False
        self.current_site = ""
        self.site_configs = {}
        self.setup_logging()
        self.setup_driver()
        self.setup_site_configs()
        
        os.makedirs('tmp_jobs', exist_ok=True)
    
    def setup_driver(self):
        """تنظیمات WebDriver"""
        try:
            options = webdriver.ChromeOptions() # type: ignore
            options.add_argument('--no-sandbox')
            options.add_argument('--disable-dev-shm-usage')
            options.add_argument('--disable-gpu')
            options.add_argument('--window-size=1920,1080')
            # options.add_argument('--headless')  # Activate if needed.
            
            self.driver = webdriver.Chrome(options=options) # type: ignore
            self.driver.implicitly_wait(5)
            
            logging.info("Driver launched")
            
        except Exception as e:
            logging.error(f"❌ Error in driver setup: {e}")
            raise
    
    def setup_site_configs(self):
        """Precise configuration for 7 target sites"""
        self.site_configs = {
            'dokhanmarket': {
               
                'name': 'Dokhan Market',
                'base_urls': ['https://dokhanmarket3.com', 'http://dokhanmarket3.com'],
                'category_selectors': [
                    'a[href*="category"]',
                    '.menu-link',
                    'nav a',
                    '.product-category a'
                ],
                'product_selectors': [
                    '.product-card',
                    '.product',
                    '.product-item',
                    '.goods-item'
                ],
                'name_selectors': [
                    '.product-card_link',
                    '.product-title',
                    'h3',
                    'h2',
                    '.product-name'
                ],
                'price_selectors': [
                    '.product-card_price',
                    '.price',
                    '.woocommerce-Price-amount',
                    '.amount',
                    'bdi'
                ],
                'pagination_selectors': [
                    'a[href*="page="]',
                    '.next',
                    '.pagination-next',
                    'a.next'
                ],
                'next_text': ['بعدی', 'next', '→'],
                'category_keywords': ['category', 'cat', 'product-category', 'shop']
            },
            'tajvape': {
              
                'name': 'Tajvape',
                'base_urls': ['https://tajvape12.com', 'http://tajvape12.com'],
                'category_selectors': [
                    '.dropdown-toggle.menu-link',
                    '.menu-link',
                    'nav a',
                    '.product-category a',
                    'a[href*="product-category"]'
                ],
                'product_selectors': [
                    'ul.products columns-4',
                    'li.col-md-3 col-6 mini-product-con type-product',
                    '.product-link',
                    '.col-md-3 col-6 mini-product-con type-product',
                    
                    'div.woocommerce shadow-box prblur mini-product product-112542 prod-variable',
                    '.product',
                    '.product-item',
                    '.woocommerce-product',
                    'li.product'
                ],
                'name_selectors': [
                    '.product-title',
                    'h2',
                    'h3',
                    '.woocommerce-loop-product__title'
                ],
                'price_selectors': [
                    '.woocommerce-Price-amount',
                    '.price',
                    '.amount',
                    'bdi'
                ],
                'pagination_selectors': [
                    '.next.page-numbers',
                    '.pagination a',
                    'a.next',
                    'a[href*="page/"]'
                ],
                'next_text': ['→', 'next', 'بعدی'],
                'category_keywords': ['product-category', 'category', 'e-juice', 'vape']
            },
            'vapoursdaily': {
              
                'name': 'Vapours Daily',
                'base_urls': ['https://vapoursdaily14.com', 'http://vapoursdaily14.com'],
                'category_selectors': [
                    '.menu-item a',
                    'nav a',
                    '.product-category a',
                    'a[href*="category"]'
                ],
                'product_selectors': [
                    '.product',
                    '.product-item',
                    '.woocommerce-product',
                    '.goods-item'
                ],
                'name_selectors': [
                    '.product-tittle',
                    'h3',
                    'h2',
                    '.product-name'
                ],
                'price_selectors': [
                    '.woocommerce-Price-amount',
                    '.price',
                    '.amount',
                    'bdi'
                ],
                'pagination_selectors': [
                    '.next.page-numbers',
                    '.pagination a',
                    'a.next'
                ],
                'next_text': ['←', 'next', 'بعدی'],
                'category_keywords': ['product-category', 'category', 'vape']
            },
            'smokcenter': {
              
                'name': 'Smok Center',
                'base_urls': ['https://smokcenter16.com', 'http://smokcenter16.com'],
                'category_selectors': [
                    'spen.elementor-icon-list-icon',
                    'spen.elementor-icon-list-text',
                    'li.elementor-icon-list-item',
                    'div.elementor-icon-wrapper',
                    '.elementor-icon',
                    '.e-n-tab-title-text',
                    'e-n-tab-title-3544064532.',
                    '.e-n-tab-title',
                    
                    '.wd-nav-products-cats a',
                    'nav a',
                    '.category-item a',
                    'a[href*="category"]'
                ],
                'product_selectors': [
                    '.product',
                    '.product-item',
                    '.wd-entities-title',
                    '.goods-item'
                ],
                'name_selectors': [
                    '.wd-entities-title',
                    'h3',
                    'h2',
                    '.product-title'
                ],
                'price_selectors': [
                    '.woocommerce-Price-amount',
                    '.price',
                    'ins .amount',
                    '.amount',
                    'bdi'
                ],
                'pagination_selectors': [
                    '.load-more-label',
                    '.next',
                    '.pagination a',
                    'a[href*="page"]'
                ],
                'next_text': ['بارگیری بیشتر محصولات', 'next', 'بعدی'],
                'category_keywords': ['shop', 'category', 'ejuice']
            },
            'digizima': {
              
                'name': 'Digi Zima',
                'base_urls': ['https://digizima19.com', 'http://digizima19.com'],
                'category_selectors': [
                    '.menu-item a',
                    'nav a',
                    '.product-category a',
                    'a[href*="category"]'
                ],
                'product_selectors': [
                    '.product',
                    '.product-item',
                    '.wd-entities-title',
                    '.goods-item'
                ],
                'name_selectors': [
                    '.wd-entities-title', 
                    'h3',
                    'h2',
                    '.product-name'
                ],
                'price_selectors': [
                    '.woocommerce-Price-amount',
                    '.price',
                    '.amount',
                    'bdi'
                ],
                'pagination_selectors': [
                    '.next.page-numbers',
                    '.pagination a',
                    'a.next'
                ],
                'next_text': ['→', 'next', 'بعدی'],
                'category_keywords': ['product-category', 'category', 'vape']
            },
            'digighelioon': {
               
                'name': 'Digi Ghelioon',
                'base_urls': ['https://digighelioon.com', 'http://digighelioon.com'],
                'category_selectors': [
                    'a.active',
                    '.menu-item a',
                    'nav a',
                    'a[href*="hookah-components"]',
                    'a[href*="category"]'
                ],
                'product_selectors': [
                    '.product',
                    '.product-item',
                    '.product-card',
                    '.goods-item'
                ],
                'name_selectors': [
                    '.product-name',
                    'h3',
                    'h2',
                    '.product-title'
                ],
                'price_selectors': [
                    '.woocommerce-Price-amount',
                    '.price',
                    '.amount',
                    'bdi'
                ],
                'pagination_selectors': [
                    '.next',
                    '.pagination a',
                    'a[href*="page"]'
                ],
                'next_text': ['بعدی', 'next', '→'],
                'category_keywords': ['product-category', 'category', 'hookah-components']
            },
            'vape60': {
                
                'name': 'Vape 60',
                'base_urls': ['https://vape60shop22.com', 'http://vape60shop22.com'],
                'category_selectors': [
                    '.menu-item a',
                    'nav a',
                    '.product-category a',
                    'a[href*="category"]'
                ],
                'product_selectors': [
                    '.product',
                    '.product-item',
                    '.woocommerce-product',
                    '.goods-item'
                ],
                'name_selectors': [
                    '.woocommerce-loop-product__title',
                    'h2',
                    'h3',
                    'b'
                ],
                'price_selectors': [
                    '.woocommerce-Price-amount',
                    '.price',
                    '.amount',
                    'bdi'
                ],
                'pagination_selectors': [
                    '.next.page-numbers',
                    '.pagination a',
                    'a.next'
                ],
                'next_text': ['←', 'next', 'بعدی'],
                'category_keywords': ['product-category', 'category', 'podsystem']
            }
        }
    
    def identify_site(self, url):
        """شناسایی هوشمند سایت بر اساس URL و محتوا"""
        logging.info(f"🔍 شناسایی سایت برای: {url}")
        
        # Identification by URL
        for site_id, config in self.site_configs.items():
            for base_url in config['base_urls']:
                if base_url in url:
                    logging.info(f"✅ سایت شناسایی شد: {config['name']}")
                    return site_id
        
        # Identification based on page content
        try:
            self.driver.get(url)
            time.sleep(3)
            page_source = self.driver.page_source
            title = self.driver.title.lower()
            
            if 'dokhan' in title or 'دخان' in page_source:
                return 'dokhanmarket'
            elif 'tajvape' in title or 'tajvape' in page_source:
                return 'tajvape'
            elif 'vapoursdaily' in title or 'vapours' in page_source:
                return 'vapoursdaily'
            elif 'smokcenter' in title or 'smok' in page_source:
                return 'smokcenter'
            elif 'digizima' in title or 'زیما' in page_source:
                return 'digizima'
            elif 'digighelioon' in title or 'قلیون' in page_source:
                return 'digighelioon'
            elif 'vape60' in title or 'vape60' in page_source:
                return 'vape60'
            else:
                logging.warning("⚠️ Unknown site, using generic configuration")
                return 'tajvape'
                
        except Exception as e:
            logging.error(f"Error in site identification: {e}")
            return 'tajvape'
    
    def setup_logging(self):
        """Reporting System Settings"""
        logging.basicConfig(
            level=logging.INFO,
            format='%(asctime)s - %(levelname)s - %(message)s',
            handlers=[
                logging.FileHandler(f'tmp_jobs/{self.job_id}_log.txt', encoding='utf-8'),
                logging.StreamHandler()
            ]
        )
    
    def update_status(self, message, page=1, total_pages=1, products_found=0, current_site=""):
        """آپدیت وضعیت"""
        status = {
            'job_id': self.job_id,
            'status': message,
            'page': page,
            'total_pages': total_pages,
            'products_count': products_found,
            'total_products': len(self.products_data),
            'current_site': current_site,
            'timestamp': datetime.now().isoformat()
        }
        
        try:
            with open(f'tmp_jobs/{self.job_id}_status.json', 'w', encoding='utf-8') as f:
                json.dump(status, f, ensure_ascii=False, indent=2)
        except Exception as e:
            print(f"Error saving status: {e}")
    
    def get_categories(self, url, site_id):
        """Get categories for a specific site"""
        self.update_status("دریافت دسته‌بندی‌ها", current_site=site_id)
        logging.info(f"🔍 Get categories from: {url} for the site{site_id}")
        
        try:
            self.driver.get(url)
            time.sleep(4)
            
            categories = []
            config = self.site_configs[site_id]
            seen_urls = set()
            
            # Method 1: Using site-specific selectors
            for selector in config['category_selectors']:
                try:
                    elements = self.driver.find_elements(By.CSS_SELECTOR, selector)
                    if elements:
                        logging.info(f"🎯 {len(elements)} Element with selector {selector}")
                        
                        for element in elements:
                            try:
                                href = element.get_attribute('href')
                                text = element.text.strip()
                                
                                if href and href not in seen_urls and text and 2 < len(text) < 100:
                                    if self.is_valid_category(href, text, site_id):
                                        categories.append({
                                            'name': text,
                                            'url': href,
                                            'site': site_id,
                                            'site_name': config['name']
                                        })
                                        seen_urls.add(href)
                                        logging.info(f"📁 Category: {text}")
                            except Exception as e:
                                logging.debug(f"Error processing element: {e}")
                                continue
                        
                        if len(categories) >= 10:
                            break
                except Exception as e:
                    logging.debug(f"Error in the selector{selector}: {e}")
                    continue
            
            # Method 2: Manually searching the menus
            if len(categories) < 3:
                categories.extend(self.find_categories_manually(site_id))
            
            # Remove duplicates
            unique_categories = []
            seen_names = set()
            for cat in categories:
                if cat['name'] not in seen_names:
                    unique_categories.append(cat)
                    seen_names.add(cat['name'])
            
            if not unique_categories:
                unique_categories.append({
                    'name': 'Main Products',
                    'url': url,
                    'site': site_id,
                    'site_name': config['name']
                })
            
            logging.info(f"📂 {len(unique_categories)} دسته‌بندی برای {site_id} یافت شد")
            return unique_categories[:12]  #Up to 12 categories
            
        except Exception as e:
            logging.error(f"Error getting categories for{site_id}: {e}")
            return [{
                'name': 'محصولات',
                'url': url,
                'site': site_id,
                'site_name': self.site_configs[site_id]['name']
            }]
    
    def find_categories_manually(self, site_id):
        """Manual search for categories"""
        categories = []
        try:
            # Search in different menus
            menu_selectors = ['nav', '.menu', '.navigation', '.main-menu', '.categories']
            
            for selector in menu_selectors:
                try:
                    menus = self.driver.find_elements(By.CSS_SELECTOR, selector)
                    for menu in menus:
                        links = menu.find_elements(By.TAG_NAME, 'a')
                        for link in links:
                            try:
                                href = link.get_attribute('href')
                                text = link.text.strip()
                                if href and text and len(text) > 2 and self.is_valid_category(href, text, site_id):
                                    categories.append({
                                        'name': text,
                                        'url': href,
                                        'site': site_id,
                                        'site_name': self.site_configs[site_id]['name']
                                    })
                            except:
                                continue
                except:
                    continue
        except:
            pass
        
        return categories
    
    def is_valid_category(self, href, text, site_id):
        """Checking the validity of the category"""
        if not href or not text:
            return False
        
        text_lower = text.lower()
        href_lower = href.lower()
        
        #forbidden words
        exclude_words = [
            'home', 'main', 'صفحه اصلی', 'contact', 'تماس', 'about', 'درباره',
            'blog', 'بلاگ', 'account', 'حساب', 'cart', 'سبد', 'checkout', 'پرداخت',
            'search', 'جستجو', 'login', 'ورود', 'register', 'ثبت نام','اسموک سنتر TV','برسی اصالت محصول'
        ]
        
        if any(word in text_lower for word in exclude_words):
            return False
        
        if any(word in href_lower for word in exclude_words):
            return False
        
        # Site-specific filters
        config = self.site_configs[site_id]
        if any(keyword in href_lower for keyword in config['category_keywords']):
            return True
        
        # General filter
        category_indicators = ['category', 'cat', 'product', 'shop', 'محصول', 'دسته']
        if any(indicator in href_lower for indicator in category_indicators):
            return True
        
        return len(text) > 2 and len(text) < 50
    
    def scrape_category_pages(self, category_url, category_name, site_id):
        """Scrape all pages of a category - **Final version with a click**"""
        logging.info(f"🔄 Start deep scraping for: {category_name}")
        
        all_products = []
        current_page = 1
        max_pages = 50
        consecutive_empty_pages = 0
        max_consecutive_empty = 1
        
        #Loading the first page
        self.driver.get(category_url)
        time.sleep(3)
        
        while current_page <= max_pages and self.is_running and consecutive_empty_pages < max_consecutive_empty:
            logging.info(f"📄 صفحه {current_page} از {category_name}")
            self.update_status(f"صفحه {current_page} از {category_name}", current_page, max_pages, len(all_products), site_id)
            
            try:
                # Scrap products from the current page
                page_products = self.scrape_products_from_page(category_name, site_id)
                
                if page_products:
                    # Filter duplicate products
                    new_products = []
                    for product in page_products:
                        if not any(p['name'] == product['name'] and p['price'] == product['price'] 
                                for p in all_products):
                            new_products.append(product)
                    
                    if new_products:
                        all_products.extend(new_products)
                        logging.info(f"✅ {len(new_products)} New product from the page{current_page}")
                        consecutive_empty_pages = 0  #Reset the counter
                    else:
                        logging.info(f"🔄 All duplicate products, page{current_page}")
                        consecutive_empty_pages += 1
                else:
                    logging.warning(f"⚠️No products on the page.{current_page}")
                    consecutive_empty_pages += 1
                
                # If 2 consecutive pages are blank/duplicate, stop.
                if consecutive_empty_pages >= max_consecutive_empty:
                    logging.info(f"🚫 {max_consecutive_empty} Blank page after page - Stop")
                    break
                
                # Try going to the next page.
                if current_page < max_pages:
                    if self.has_next_page_improved(site_id):
                        if self.click_next_page(site_id):
                            current_page += 1
                            time.sleep(2)
                        else:
                            # If it can't click, go with the direct URL.
                            logging.info("🔄 Use direct URL for next page")
                            next_url = self.get_page_url(category_url, current_page + 1, site_id)
                            self.driver.get(next_url)
                            time.sleep(3)
                            current_page += 1
                    else:
                        logging.info("🏁 There is no next page - End of category")
                        break
                else:
                    logging.info("🏁 We have reached the maximum number of pages allowed.")
                    break
                    
            except Exception as e:
                logging.error(f"❌ Error on the page{current_page}: {e}")
                consecutive_empty_pages += 1
                
                # Try going to the next page with the direct URL.
                try:
                    next_url = self.get_page_url(category_url, current_page + 1, site_id)
                    self.driver.get(next_url)
                    time.sleep(3)
                    current_page += 1
                except:
                    break
        
        logging.info(f"🎉 Completion {category_name}: {len(all_products)} product of{current_page} page")
        return all_products
                
    def get_page_url(self, base_url, page_number, site_id):
        """Page URL Builder - **Supports all formats**"""
        if page_number == 1:
            return base_url
            
        # Delete existing pagination parameters
        base_clean = re.sub(r'[?&](page|paged)=\d+', '', base_url)
        base_clean = re.sub(r'/page/\d+', '', base_clean)
        base_clean = re.sub(r'/product-page/\d+', '', base_clean)
            
        # Add a new page based on site type
        if site_id in ['tajvape', 'vapoursdaily', 'digizima']:
            # Format: /page/2/
            return f"{base_clean}/page/{page_number}/"
        elif site_id in ['smokcenter', 'vape60']:
            # Format: ?page=2
            separator = '?' if '?' not in base_clean else '&'
            return f"{base_clean}{separator}page={page_number}"
        elif site_id in ['dokhanmarket', 'digighelioon']:
                # Format: /product-page/2/
            return f"{base_clean}/product-page/{page_number}/"
        else:
                #Default format
            separator = '?' if '?' not in base_clean else '&'
            return f"{base_clean}{separator}page={page_number}"
        
    def has_next_page_improved(self, site_id):
        """Checking for the existence of the next page - **Super Advanced Version**"""
        config = self.site_configs[site_id]
        current_url = self.driver.current_url
        
        logging.info(f"🔍 Search the next page for{config['name']}")
        
        # Method 1: Search for "Next" buttons with different selectors
        next_selectors = [
            'a.next', '.next', '.pagination-next', 
            '.page-numbers.next', '.next.page-numbers',
            'a[rel="next"]', '.next-page', '.pagination .next',
            '.woocommerce-pagination .next', '.nav-next',
            'a:contains("بعدی")', 'a:contains("next")',
            'button.next', '.load-more', '.pagination-next a'
        ]
        
        for selector in next_selectors:
            try:
                next_elements = self.driver.find_elements(By.CSS_SELECTOR, selector)
                for element in next_elements:
                    try:
                        if element.is_displayed() and element.is_enabled():
                            text = element.text.lower().strip()
                            href = element.get_attribute('href') or ''
                            
                            # کلمات کلیدی برای صفحه بعد
                            next_keywords = ['next', 'بعدی', '→', '»', '>', 'load more', 'more']
                            prev_keywords = ['قبلی', 'قبل', '←', '«', '<', 'previous']
                            
                            if (any(keyword in text for keyword in next_keywords) and 
                                not any(keyword in text for keyword in prev_keywords)):
                                logging.info(f"🎯 The next page was found with the selector: {selector}")
                                return True
                    except:
                        continue
            except:
                continue
        
        # Method 2: Search the entire page for pagination links
        try:
           # All possible links for pagination
            all_links = self.driver.find_elements(By.CSS_SELECTOR, 
                'a[href*="page"], a[href*="paged"], [class*="page"], [class*="pagination"] a, .page-numbers a, .pagination a, .page-links a')
            
            current_page = self.get_current_page_number(current_url)
            
            for link in all_links:
                try:
                    if not link.is_displayed():
                        continue
                        
                    link_text = link.text.strip()
                    href = link.get_attribute('href')
                    
                    if not href:
                        continue
                    
                   # If the link is the next page number
                    if link_text.isdigit():
                        link_page = int(link_text)
                        if link_page == current_page + 1:
                            logging.info(f"🔢 Next page found: page{link_page}")
                            return True
                    
                    # If the link contains words from the next page
                    text_lower = link_text.lower()
                    if any(word in text_lower for word in ['next', 'بعدی', '→', '»', '>']):
                        if not any(word in text_lower for word in ['قبلی', 'قبل', '←']):
                            logging.info(f"📖 Next page with text:{link_text}")
                            return True
                            
                except:
                    continue
        except Exception as e:
            logging.debug(f"Error searching for links: {e}")
        
        # Method 3: Search with XPath for specific text
        try:
            next_texts = ['بعدی', 'next', '→', '»', '>', 'Load more', 'More products']
            for text in next_texts:
                try:
                    elements = self.driver.find_elements(By.XPATH, f"//*[contains(text(), '{text}')]")
                    for element in elements:
                        try:
                            if element.is_displayed() and element.is_enabled():
                                # Check that the element is really for the next page
                                parent = element.find_element(By.XPATH, './..')
                                if parent.tag_name == 'a' or parent.get_attribute('onclick'):
                                    logging.info(f"🔍 Next page with XPath: {text}")
                                    return True
                        except:
                            continue
                except:
                    continue
        except Exception as e:
            logging.debug(f"Error in XPath search: {e}")
        
        # Method 4: Check for changes in URL after click (for Load More)
        try:
            # Finding elements that may be Load More
            buttons = self.driver.find_elements(By.CSS_SELECTOR, 
                'button, [onclick], [class*="load"], [class*="more"]')
            
            for button in buttons:
                try:
                    if button.is_displayed() and button.is_enabled():
                        text = button.text.lower()
                        if any(word in text for word in ['more', 'load', 'بارگیری', 'بیشتر']):
                            logging.info(f"🔄 Load More button found: {text}")
                            return True
                except:
                    continue
        except:
            pass
        
        logging.info("❌ No next page found")
        return False
    
    def click_next_page(self, site_id):
        """Click on the next page - **New function**"""
        config = self.site_configs[site_id]
        current_url = self.driver.current_url
        
        logging.info("🖱️ تلاش برای کلیک روی صفحه بعد... ")
        
        # Method 1: Clicking "Next" buttons with different selectors
        next_selectors = [
            'a.next', '.next', '.pagination-next', 
            '.page-numbers.next', '.next.page-numbers',
            'a[rel="next"]', '.next-page', '.pagination .next',
            '.woocommerce-pagination .next', '.nav-next'
        ]
        
        for selector in next_selectors:
            try:
                next_buttons = self.driver.find_elements(By.CSS_SELECTOR, selector)
                for button in next_buttons:
                    try:
                        if button.is_displayed() and button.is_enabled():
                            logging.info(f"✅ Click on the next page with the selector: {selector}")
                            self.driver.execute_script("arguments[0].click();", button)
                            time.sleep(3)
                            return True
                    except Exception as e:
                        logging.debug(f"Error when clicking with selector{selector}: {e}")
                        continue
            except Exception as e:
                logging.debug(f"Error finding selector{selector}: {e}")
                continue
        
        # Method 2: Click on the next page number
        try:
            current_page = self.get_current_page_number(current_url)
            page_links = self.driver.find_elements(By.CSS_SELECTOR, 
                '.page-numbers a, .pagination a, a.page-numbers, .page-links a')
            
            for link in page_links:
                try:
                    if link.is_displayed() and link.is_enabled():
                        link_text = link.text.strip()
                        if link_text.isdigit():
                            link_page = int(link_text)
                            if link_page == current_page + 1:
                                logging.info(f"🔢 Click on the page{link_page}")
                                self.driver.execute_script("arguments[0].click();", link)
                                time.sleep(3)
                                return True
                except:
                    continue
        except Exception as e:
            logging.debug(f"Error clicking on page numbers:{e}")
        
        # Method 3: Click on "Next" texts with XPath
        try:
            next_texts = ['بعدی', 'next', '→', '»', '>']
            for text in next_texts:
                try:
                    elements = self.driver.find_elements(By.XPATH, f"//*[contains(text(), '{text}')]")
                    for element in elements:
                        try:
                            if element.is_displayed() and element.is_enabled():
                                # Check that the element is really for the next page
                                element_text = element.text.lower()
                                if not any(word in element_text for word in ['قبلی', 'قبل', '←', '«']):
                                    logging.info(f"📖 Click on: {text}")
                                    self.driver.execute_script("arguments[0].click();", element)
                                    time.sleep(3)
                                    return True
                        except:
                            continue
                except:
                    continue
        except Exception as e:
            logging.debug(f"Error in XPath click: {e}")
        
        # Method 4: Clicking on the Load More buttons
        try:
            load_more_selectors = [
                'button.load-more', '.load-more', '[class*="load-more"]',
                '.load-more-products', '.ajax-load-more',
                'button:contains("Load more")', 'button:contains("load more")'
            ]
            
            for selector in load_more_selectors:
                try:
                    buttons = self.driver.find_elements(By.CSS_SELECTOR, selector)
                    for button in buttons:
                        try:
                            if button.is_displayed() and button.is_enabled():
                                logging.info(f"🔄 کلیک روی Load More: {selector}")
                                self.driver.execute_script("arguments[0].click();", button)
                                time.sleep(4) # More time to load new products
                                return True
                        except:
                            continue
                except:
                    continue
        except Exception as e:
            logging.debug(f"خطا در کلیک Load More: {e}")
        
        logging.warning("❌ Could not click on the next page")
        return False
        
    def get_current_page_number(self, url):
            """Get current page number from URL - **Modified**"""
            try:
                patterns = [
                    r'/page/(\d+)/',
                    r'[?&]page=(\d+)',
                    r'/product-page/(\d+)/',
                    r'[?&]paged=(\d+)',
                    r'/page-(\d+)/',
                    r'/page(\d+)/'
                ]
                
                for pattern in patterns:
                    match = re.search(pattern, url)
                    if match:
                        page_num = int(match.group(1))
                        logging.info(f"📖 Current page number: {page_num}")
                        return page_num
                
                # If the page number is not found, it is probably the first page.
                return 1
            except:
                return 1
    
    def scrape_products_from_page(self, category_name, site_id):
        """Scrap products with duplicate filter"""
        products = []
        config = self.site_configs[site_id]
        
        for selector in config['product_selectors']:
            try:
                elements = self.driver.find_elements(By.CSS_SELECTOR, selector)
                if elements:
                    logging.info(f"🎯 {len(elements)} element with{selector}")
                    
                    for element in elements:
                        try:
                            if not self.is_running:
                                break
                                
                            product = self.extract_product_data(element, category_name, site_id)
                            if product and self.is_valid_product(product):
                                # Check for duplicates on the same page
                                if not any(p['name'] == product['name'] and p['price'] == product['price'] 
                                        for p in products):
                                    products.append(product)
                        except Exception as e:
                            continue
                    
                    if products:
                        break
            except:
                continue
        
        return products
    
    def is_duplicate_product(self, new_product, existing_products):
        """Checking for product non-duplicateness"""
        for existing in existing_products:
            if (existing['name'] == new_product['name'] and 
                existing['price'] == new_product['price'] and
                existing['site'] == new_product['site']):
                return True
        return False
    
    def extract_product_data(self, element, category_name, site_id):
        """Product Information Extraction - **Modified**"""
        try:
            full_text = element.text.strip()
            if len(full_text) < 10:  # Reduce the minimum text length
                return None
            
            # Name extraction
            name = self.extract_product_name(element, site_id)
            if not name or len(name) < 2:  # کاهش حداقل طول نام
                lines = [line.strip() for line in full_text.split('\n') if line.strip()]
                name = lines[0] if lines else "محصول ناشناخته"
            
            # Price extraction
            price = self.extract_product_price(element, site_id)
            if not price:
                price = self.extract_price_from_text(full_text)
            
            if not price:
                return None
            
            # URL extraction
            url = self.extract_product_url(element, site_id)
            
            # SKU extraction
            sku = self.extract_sku(element, full_text, site_id)
            
            product_data = {
                'name': name[:200],  # Increase name length
                'price': price,
                'categories': category_name,
                'site': self.site_configs[site_id]['name'],
                'site_id': site_id,
                'type': 'product',
                'variation': 'standard',
                'sku': sku,
                'description': full_text[:300],  # Increasing the length of the description
                'url': url,
                'grouped_products': '',
                'scraped_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            }
            
            return product_data
            
        except Exception as e:
            logging.debug(f"Error extracting product: {e}")
            return None
    
    def extract_product_name(self, element, site_id):
        """Product Name Extraction - **Modified**"""
        config = self.site_configs[site_id]
        
        for selector in config['name_selectors']:
            try:
                if selector in ['h2', 'h3', 'h4', 'b', 'strong']:
                    # If the selector is an HTML tag
                    if element.tag_name == selector:
                        name = element.text.strip()
                        if name and len(name) > 1:
                            return name
                    # Or finding in children
                    try:
                        name_elems = element.find_elements(By.TAG_NAME, selector)
                        for name_elem in name_elems:
                            name = name_elem.text.strip()
                            if name and len(name) > 1:
                                return name
                    except:
                        continue
                else:
                    # Normal CSS selector
                    name_elems = element.find_elements(By.CSS_SELECTOR, selector)
                    for name_elem in name_elems:
                        name = name_elem.text.strip()
                        if name and len(name) > 1:
                            return name
            except:
                continue
        
        return None
    
    def extract_product_price(self, element, site_id):
        """Product Price Extraction - **Modified**"""
        config = self.site_configs[site_id]
        
        for selector in config['price_selectors']:
            try:
                price_elems = element.find_elements(By.CSS_SELECTOR, selector)
                for price_elem in price_elems:
                    try:
                        price_text = price_elem.text.strip()
                        price = self.extract_price_from_text(price_text)
                        if price:
                            return price
                    except:
                        continue
            except:
                continue
        
        return None
    
    def extract_price_from_text(self, text):
        """Extracting Price from Text - **Modified**"""
        try:
            # Clear text and preserve numbers and separators
            clean_text = re.sub(r'[^\d,\.\s]', '', text.strip())
            clean_text = re.sub(r'\s+', ' ', clean_text)
            
            # Different price patterns
            patterns = [
                r'(\d{1,3}(?:,\d{3})*(?:\.\d+)?)',  # format 1,000,000
                r'(\d{1,3}(?:\.\d{3})*(?:,\d+)?)',  # format 1,000,000
                r'(\d+)'  # Just numbers
            ]
            
            for pattern in patterns:
                matches = re.findall(pattern, clean_text)
                for match in matches:
                    try:
                        #Remove separators and convert to numbers
                        price_str = re.sub(r'[^\d]', '', match)
                        if price_str.isdigit():
                            price = int(price_str)
                            # Reasonable price range for vape products
                            if 1000 <= price <= 50000000:
                                return str(price)
                    except:
                        continue
        except:
            pass
        
        return None
    
    def extract_product_url(self, element, site_id):
        """Product URL Extraction"""
        try:
            # If the element itself is a link
            if element.tag_name == 'a':
                href = element.get_attribute('href')
                if href and 'http' in href:
                    return href
            
            # Search for links in children
            links = element.find_elements(By.TAG_NAME, 'a')
            for link in links:
                href = link.get_attribute('href')
                if href and 'http' in href:
                    return href
            
            return ""
        except:
            return ""
    
    def extract_sku(self, element, text, site_id):
        """Product SKU Extraction"""
        try:
            sku_patterns = [
                r'SKU:\s*([A-Za-z0-9-]+)',
                r'کد:\s*([A-Za-z0-9-]+)',
                r'شناسه:\s*([A-Za-z0-9-]+)',
                r'([A-Z]{2,3}\d{3,})',
                r'کد محصول:\s*([^\s]+)'
            ]
            
            for pattern in sku_patterns:
                matches = re.findall(pattern, text)
                if matches:
                    return matches[0]
        except:
            pass
        
        return ""
    
    def is_valid_product(self, product):
        """Product Validation Check - **Modified**"""
        if not product.get('name') or len(product['name']) < 2:
            return False
        
        if not product.get('price') or not product['price'].isdigit():
            return False
        
        price_num = int(product['price'])
        if price_num < 500 or price_num > 100000000:  # Extending the price range
            return False
        
        return True
    
    def alternative_scraping_methods(self, category_name, site_id):
        """Alternative Methods for Scraping - **Modified**"""
        products = []
        
        try:
            # Search for elements containing prices
            price_indicators = ['تومان', 'ریال', 'price', 'قیمت', 'خرید']
            for indicator in price_indicators:
                try:
                    elements = self.driver.find_elements(By.XPATH, f'//*[contains(text(), "{indicator}")]')
                    for element in elements[:50]:  # Increasing the number of elements
                        try:
                            # Finding the parent that likely contains product information
                            parent = element.find_element(By.XPATH, './ancestor::*[position()<5]')
                            text = parent.text.strip()
                            if len(text) > 30 and self.looks_like_product(text):
                                product = self.create_product_from_text(text, category_name, site_id)
                                if product and self.is_valid_product(product) and not self.is_duplicate_product(product, products):
                                    products.append(product)
                        except:
                            continue
                except:
                    continue
        except:
            pass
        
        return products
    
    def create_product_from_text(self, text, category_name, site_id):
        """Create a product from text"""
        try:
            lines = [line.strip() for line in text.split('\n') if line.strip() and len(line.strip()) > 2]
            if not lines:
                return None
            
            # Find the name (first sensible line)
            name = lines[0]
            for line in lines:
                if len(line) > 5 and not any(indicator in line.lower() for indicator in ['تومان', 'ریال', 'قیمت', 'price', 'خرید']):
                    name = line
                    break
            
            # Price extraction
            price = self.extract_price_from_text(text)
            if not price:
                return None
            
            return {
                'name': name[:200],
                'price': price,
                'categories': category_name,
                'site': self.site_configs[site_id]['name'],
                'site_id': site_id,
                'type': 'product',
                'variation': 'standard',
                'sku': '',
                'description': text[:300],
                'url': '',
                'grouped_products': '',
                'scraped_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            }
        except:
            return None
    
    def looks_like_product(self, text):
        """Checking if the text is similar to the product - **Modified**"""
        must_have = ['تومان', 'ریال', 'price']
        nice_to_have = ['قیمت', 'خرید', 'جویس', 'پاد', 'ویپ', 'کویل', 'سیستم', 'محصول', 'product', 'vape']
        
        text_lower = text.lower()
        
        if not any(indicator in text_lower for indicator in must_have):
            return False
        
        if any(indicator in text_lower for indicator in nice_to_have):
            return True
        
        return len(text) > 50  # Reduce the minimum length
    
    def scrape_all_sites(self):
        """Scrap all 7 target sites"""
        target_sites = [
            "https://vape60shop22.com",
            "https://tajvape12.com", 
            "https://vapoursdaily14.com",
            "https://digizima19.com",
            "https://smokcenter16.com",
            "https://digighelioon.com",
            "https://dokhanmarket3.com"
        ]
        
        return self.scrape_multiple_sites(target_sites)
    
    def scrape_multiple_sites(self, site_urls):
        """Multiple Site Scraping - **Final Fix**"""
        self.is_running = True
        total_results = []
        
        try:
            for i, site_url in enumerate(site_urls, 1):
                if not self.is_running:
                    break
                
                logging.info(f"🌐 Start scraping the site{i}/{len(site_urls)}: {site_url}")
                self.update_status(f"سایت {i}", current_site=site_url)
                
                #Site identification
                site_id = self.identify_site(site_url)
                self.current_site = site_id
                
                #Get categories
                categories = self.get_categories(site_url, site_id)
                logging.info(f"📂 {len(categories)} Categories for {site_id} found")
                
                site_products = []
                
                # Scrap each category
                for j, category in enumerate(categories, 1):
                    if not self.is_running:
                        break
                    
                    logging.info(f"🔄 دسته‌بندی {j}/{len(categories)}: {category['name']}")
                    
                    # **Main fix: Return to home page before each new category**
                    try:
                        self.driver.get(site_url)  # Back to the main page
                        time.sleep(2)
                    except:
                        pass
                    
                    # Scrape all pages in this category
                    category_products = self.scrape_category_pages(
                        category['url'], 
                        category['name'], 
                        site_id
                    )
                    
                    if category_products:
                        site_products.extend(category_products)
                        logging.info(f"✅ {len(category_products)} product of{category['name']}")
                    
                    time.sleep(2)  # Delay between categories
                    
                    # **Temporary storage after each classification**
                    self.products_data.extend(site_products)
                    self.save_progress()
                    
                    #**Status update to show progress**
                    self.update_status(
                        f"دسته‌بندی {j}/{len(categories)} از سایت {i}", 
                        current_site=site_id
                    )
                
                # **Final storage of products on this site**
                if site_products:
                    total_results.append({
                        'site': site_id,
                        'site_name': self.site_configs[site_id]['name'],
                        'url': site_url,
                        'categories_count': len(categories),
                        'products_count': len(site_products),
                        'status': 'success'
                    })
                    
                    logging.info(f"✅ اتمام سایت {site_id}: {len(site_products)} محصول")
                else:
                    logging.warning(f"⚠️ هیچ محصولی از سایت {site_id} یافت نشد")
                
                time.sleep(3)  # Latency between sites
            
            # **Final storage of all products**
            excel_file = self.save_to_excel()
            
            final_result = {
                'success': True,
                'job_id': self.job_id,
                'total_products': len(self.products_data),
                'sites_scraped': len(total_results),
                'excel_file': excel_file,
                'site_results': total_results,
                'message': f'تعداد {len(self.products_data)} product of{len(total_results)} site found'
            }
            
            logging.info(f"🎉 Complete scrap completion: {final_result}")
            return final_result
            
        except Exception as e:
            error_msg = f"error: {str(e)}"
            logging.error(f"❌ {error_msg}")
            return {
                'success': False,
                'error': error_msg,
                'job_id': self.job_id
            }
        finally:
            self.is_running = False
    
    def save_progress(self):
        """ذخیره پیشرفت"""
        try:
            progress_data = {
                'job_id': self.job_id,
                'products': self.products_data,
                'total_products': len(self.products_data),
                'current_site': self.current_site,
                'timestamp': datetime.now().isoformat()
            }
            
            with open(f'tmp_jobs/{self.job_id}.json', 'w', encoding='utf-8') as f:
                json.dump(progress_data, f, ensure_ascii=False, indent=2)
                
        except Exception as e:
            logging.error(f"Error saving progress: {e}")
    
    def save_to_excel(self):
        """Save to Excel with Duplicates Removed - **Final Version**"""
        if not self.products_data:
            logging.warning("⚠️There is no data to save.")
            return None
        
        try:
            filename = f"tmp_jobs/{self.job_id}.xlsx"
            
            # Creating a DataFrame from data
            df = pd.DataFrame(self.products_data)
            
            # **Remove duplicates before saving**
            initial_count = len(df)
            
            # Remove duplicates by name, price, and site
            df = df.drop_duplicates(
                subset=['name', 'price', 'site'], 
                keep='first'
            )
            
            # Also remove exact duplicates (all fields)
            df = df.drop_duplicates(keep='first')
            
            final_count = len(df)
            duplicates_removed = initial_count - final_count
            
            logging.info(f"🧹 delete {duplicates_removed} Duplicate case of{initial_count} product")
            
           # If all data were duplicates
            if len(df) == 0:
                logging.warning("⚠️All data was duplicated - save at least one record")
               # Keep at least one record of the original data
                df = pd.DataFrame(self.products_data[:1])
            
            # Create an Excel file with formatting
            wb = Workbook()
            ws = wb.active
            ws.title = "Products"
            
           # Add headers
            headers = list(df.columns)
            ws.append(headers)
            
            # Add non-repeating data
            for _, row in df.iterrows():
                ws.append(row.tolist())
            
           # Add statistical information in a separate sheet
            stats_sheet = wb.create_sheet(title="آمار")
            stats_data = [
                ["آمار محصولات استخراج شده"],
                ["تاریخ استخراج", datetime.now().strftime('%Y-%m-%d %H:%M:%S')],
                ["تعداد کل محصولات پیدا شده", initial_count],
                ["تعداد محصولات منحصر به فرد", final_count],
                ["تعداد موارد تکراری حذف شده", duplicates_removed],
                ["تعداد سایت‌ها", len(df['site'].unique())],
                [],
                ["تعداد محصولات هر سایت:"]
            ]
            
            # Statistics of each site
            site_stats = df['site'].value_counts()
            for site, count in site_stats.items():
                stats_data.append([site, count])
            
            for row in stats_data:
                stats_sheet.append(row)
            
            # formatting
            self.apply_excel_styling(ws, len(df))
            
            # Formatting a statistics sheet
            try:
                for col in range(1, 3):
                    stats_sheet.column_dimensions[get_column_letter(col)].width = 30
                
                for row in range(1, len(stats_data) + 1):
                    for col in range(1, 3):
                        cell = stats_sheet.cell(row=row, column=col)
                        if row == 1:
                            cell.font = Font(bold=True, size=14, color="1565C0")
                        elif row <= 7:
                            cell.font = Font(bold=True, color="2E7D32")
            except:
                pass
            
            # ذخیره فایل
            wb.save(filename)
            logging.info(f"💾 Excel file saved: {filename} (with {final_count} Unique product)")
            
            #Also save a JSON file with non-duplicate data.
            unique_data = {
                'job_id': self.job_id,
                'total_products_initial': initial_count,
                'total_products_final': final_count,
                'duplicates_removed': duplicates_removed,
                'products': df.to_dict('records'),
                'timestamp': datetime.now().isoformat()
            }
            
            with open(f'tmp_jobs/{self.job_id}_unique.json', 'w', encoding='utf-8') as f:
                json.dump(unique_data, f, ensure_ascii=False, indent=2)
            
            return filename
            
        except Exception as e:
            logging.error(f"❌ Error saving Excel: {e}")
            # Simple save in case of error
            try:
                simple_filename = f"tmp_jobs/{self.job_id}_simple.xlsx"
                df = pd.DataFrame(self.products_data)
                df.to_excel(simple_filename, index=False, engine='openpyxl')
                return simple_filename
            except Exception as e2:
                logging.error(f"❌ Error in simple save: {e2}")
                return None
    
    def apply_excel_styling(self, worksheet, data_count):
        """Applying Beautiful Styles to Excel"""
        try:
            #Soft and eye-catching colors
            header_fill = PatternFill(start_color="18AAC4", end_color="18AAC4", fill_type="solid")  # Very light blue
            even_row_fill = PatternFill(start_color="C2F0FF", end_color="C2F0FF", fill_type="solid")  # Very soft gray
            odd_row_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")   # white
            price_fill = PatternFill(start_color="F0F8EB", end_color="F0F8EB", fill_type="solid")     # Very light green
            site_fill = PatternFill(start_color="F0F8EB", end_color="F0F8EB", fill_type="solid")      # Very mild orange
            
            # fonts
            header_font = Font(bold=True, color="2E4057", size=11)
            normal_font = Font(color="2D2D2D", size=10)
            price_font = Font(bold=True, color="2E8B57", size=10)
            site_font = Font(bold=True, color="2E4057", size=10)
            
            # level
            center_align = Alignment(horizontal='center', vertical='center')
            right_align = Alignment(horizontal='right', vertical='center')
            left_align = Alignment(horizontal='left', vertical='center')
            
            # Header formatting
            for col in range(1 , len(worksheet[1]) + 1):
                cell = worksheet.cell(row=1, column=col)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = center_align
            
            # Data formatting
            for row in range(2, data_count + 2):
                # Coloring rows one by one
                if row % 2 == 0:
                    row_fill = even_row_fill
                else:
                    row_fill = odd_row_fill
                
                for col in range(1, len(worksheet[1]) + 1):
                    cell = worksheet.cell(row=row, column=col)
                    cell.font = normal_font
                    cell.fill = row_fill
                    
                    header_value = worksheet.cell(row=1, column=col).value
                    
                    # Special price format
                    if header_value == 'price':
                        cell.font = price_font
                        cell.fill = price_fill
                        cell.alignment = right_align
                    #Special format for the site
                    elif header_value in ['site', 'site_id']:
                        cell.font = site_font
                        cell.fill = site_fill
                        cell.alignment = center_align
                    # Special name format
                    elif header_value == 'name':
                        cell.alignment = left_align
                    else:
                        cell.alignment = right_align
            
            #Adjusting column widths
            column_widths = {
                'name': 80,
                'price': 15,
                'categories': 25,
                'site': 20,
                'site_id': 15,
                'description': 130,
                'url': 50
            }
            
            for col in range(1, len(worksheet[1]) + 1):
                header = worksheet.cell(row=1, column=col).value
                if header in column_widths:
                    worksheet.column_dimensions[get_column_letter(col)].width = column_widths[header]
                else:
                    worksheet.column_dimensions[get_column_letter(col)].width = 15
            
            # Freeze header
            worksheet.freeze_panes = "A2"
            
            logging.info("🎨 Excel formatting applied.")
            
        except Exception as e:
            logging.warning(f"Error applying styles: {e}")
    
    def stop(self):
        """Stop Scraping"""
        self.is_running = False
    
    def close(self):
        """Close Driver"""
        if self.driver:
            try:
                self.driver.quit()
                logging.info("🔚 Driver closed")
            except:
                pass

# Main function to run
def main():
    """Main function to run the scraper"""
    scraper = AdvancedVapeScraper()
    
    try:
        result = scraper.scrape_all_sites()
        print("Results:", result)
        
        if result['success']:
            print(f"🎉 Scraping was successful!")
            print(f"📊 Number of products: {result['total_products']}")
            print(f"🌐 Number of sites: {result['sites_scraped']}")
            print(f"💾 Excel file: {result['excel_file']}")
        else:
            print(f"❌ خطا: {result['error']}")
            
    except Exception as e:
        print(f"Error in main execution: {e}")
    finally:
        scraper.close()

if __name__ == "__main__":
    main()